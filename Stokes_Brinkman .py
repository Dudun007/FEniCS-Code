#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri May 14 14:59:47 2021

@author: C00257297
"""

from fenics import *
from mshr import *
import numpy as np
import sys
import random
import math
import matplotlib.pyplot as plt

from openpyxl import load_workbook
import pandas as pd

# import the builtin time module
import time



#Initial and Boundary Condition
u_inlet = 0.000167*100          #m/sec (1 m/sec = 6000 cm/min)

#Rock and Fluid Properties
mu = Constant(0.001)            #pa-sec (1 pa-sec = 1000 cp)
poro_ref = 0.2                  #fraction
perm_ref = 150E-15               #m2 (1 mD = 1E-15 m2)

# Create a computational domain
DimX = 1
DimY = 1
domain = Rectangle(Point(0,0), Point(DimX,DimY))

# Create a subdomain for fractures (each frac is a rectangle)
def createFracSubDomain (fracdata):
  fracdomain = []

  for i in range (0, len(fracdata)):
    L = fracdata[i][0]      # the length of your rectangle
    b = fracdata[i][1]      # aperture
    a = fracdata[i][2]      # orientation angle
    Cx = fracdata[i][3]     # center point x
    Cy = fracdata[i][4]     # center point y

    # coordinate of each corner point for rectangle with the center at (0,0)
    LeftTop_x = -L/2; LeftTop_y = b/2; LeftBot_x = -L/2; LeftBot_y = -b/2
    RightBot_x = L/2; RightBot_y = -b/2; RightTop_x = L/2; RightTop_y = b/2

    #The rotated position of each corner    
    Rx1 = Cx + (LeftTop_x  * math.cos(a)) - (LeftTop_y * math.sin(a))
    Ry1 = Cy + (LeftTop_x  * math.sin(a)) + (LeftTop_y * math.cos(a))
    Rx2 = Cx + (LeftBot_x  * math.cos(a)) - (LeftBot_y * math.sin(a))
    Ry2 = Cy + (LeftBot_x  * math.sin(a)) + (LeftBot_y * math.cos(a))
    Rx3 = Cx + (RightBot_x  * math.cos(a)) - (RightBot_y * math.sin(a))
    Ry3 = Cy + (RightBot_x  * math.sin(a)) + (RightBot_y * math.cos(a))
    Rx4 = Cx + (RightTop_x  * math.cos(a)) - (RightTop_y * math.sin(a))
    Ry4 = Cy + (RightTop_x  * math.sin(a)) + (RightTop_y * math.cos(a))

    domain_vertices = [Point(Rx1,Ry1),Point(Rx2,Ry2),Point(Rx3,Ry3),Point(Rx4,Ry4),Point(Rx1,Ry1)]

    if i == 0:
      fracdomain = Polygon(domain_vertices)
    else:
      fracdomain = fracdomain + Polygon(domain_vertices)

  return fracdomain

FracWidth = 0.001
# generate fracture subdmain based on fracture data 
# fracdata = [length, aperture, angle, center_x, center_y]


fracdata = [[0.6,FracWidth, 0, 0.5, 0.5]] #Horizontal fracture
#fracdata = [[0.6,FracWidth, math.pi/2, 0.5, 0.5]] #Vertical fracture
#fracdata = [[0.6,FracWidth, 0.785,  0.5, 0.5]] #Diagonal fracture
#fracdata = [[0.0000001,FracWidth, 0, 1, 1]] #No fracture

fracdomain = createFracSubDomain (fracdata)




# Add fractures into the domain and mark them as subdomain '1'
domain.set_subdomain(1, fracdomain)
domain.set_subdomain(2, domain - fracdomain)

# Generature mesh for the combined system
mesh = generate_mesh(domain, 100)

# Create a cell list for fractures -> frac_cells
domains = mesh.domains()
subdomains = MeshFunction('size_t',mesh,2,domains)
frac_cells = SubsetIterator(subdomains, 1)

#plot(mesh)

# Define function spaces and mixed (product) space
scalar = FiniteElement("CG", mesh.ufl_cell(), 1)
vector = VectorElement("CG", mesh.ufl_cell(), 2)
W = FunctionSpace(mesh, MixedElement(vector, scalar))

# Define trial and test functions
(u, P) = TrialFunctions(W)
(tf1, tf2) = TestFunctions(W)


#Initialize Properties and Generate Cell Functions for Properties
poro = MeshFunction("double", mesh, 2)
perm = MeshFunction("double", mesh, 2)
for cell in cells(mesh):
    poro[cell] = 0.175 #poro_ref#random.gauss(poro_ref, 0.01)
    perm[cell] = perm_ref*pow(pow(poro[cell]/poro_ref,3)*pow((1-poro_ref)/(1-poro[cell]),2),6)

#update properties in fracture cells
for cell in frac_cells:
    poro[cell] = 0.9999999
    perm[cell] = perm_ref*pow(pow(poro[cell]/poro_ref,3)*pow((1-poro_ref)/(1-poro[cell]),2),6)

# Define Properties Functions and Initialize###############################
#plot(poro)
print(max(poro.array()[:]),min(poro.array()[:]))
print(max(perm.array()[:]),min(perm.array()[:]))


# Code for C++ evaluation of conductivity
property_code = """

#include <pybind11/pybind11.h>
#include <pybind11/eigen.h>
namespace py = pybind11;

#include <dolfin/function/Expression.h>
#include <dolfin/mesh/MeshFunction.h>

class Property : public dolfin::Expression
{
public:

  // Create expression with 2 components
  Property() : dolfin::Expression(2) {}

  // Function for evaluating expression on each cell
  void eval(Eigen::Ref<Eigen::VectorXd> values, Eigen::Ref<const Eigen::VectorXd> x, const ufc::cell& cell) const override
  {
    const uint cell_index = cell.index;
    values[0] = (*c0)[cell_index];
    values[1] = (*c1)[cell_index];
  }

  // The data stored in mesh functions
  std::shared_ptr<dolfin::MeshFunction<double>> c0;
  std::shared_ptr<dolfin::MeshFunction<double>> c1;

};

PYBIND11_MODULE(SIGNATURE, m)
{
  py::class_<Property, std::shared_ptr<Property>, dolfin::Expression>
    (m, "Property")
    .def(py::init<>())
    .def_readwrite("c0", &Property::c0)
    .def_readwrite("c1", &Property::c1);
}

"""

c = CompiledExpression(compile_cpp_code(property_code).Property(),
                       c0=poro, c1=perm, degree=0)

Phi = c[0]
K = c[1]

# Create functions for boundary conditions
noslip = Constant((0, 0))
inflow = Constant((u_inlet, 0))
zero   = Constant(0)

#Define boundary
def boundary1(x):
    return x[0] < DOLFIN_EPS and x[1] > DOLFIN_EPS and x[1] < DimY - DOLFIN_EPS

def boundary2(x):
    return (x[1] < DOLFIN_EPS or x[1] > DimY - DOLFIN_EPS) #or (x[0] < DOLFIN_EPS and x[1] <= 0.48) or (x[0] < DOLFIN_EPS and x[1] >= 1-0.48)

def boundary3(x):
    return x[0] > DimX - DOLFIN_EPS

# Define essential boundary
# No-slip boundary condition for velocity
bc0 = DirichletBC(W.sub(0), noslip, boundary2)
# Inflow boundary condition for velocity
bc1 = DirichletBC(W.sub(0), inflow, boundary1)
# Boundary condition for pressure at outflow
bc2 = DirichletBC(W.sub(1), zero, boundary3)
# Collect boundary conditions
bcs = [bc0, bc1, bc2]


f = Constant(0.0)

a = mu/Phi*inner(grad(u),grad(tf1))*dx + inner(grad(P),tf1)*dx + mu/K*inner(u,tf1)*dx + div(u)*tf2*dx
L = dot(f,tf2)*dx

w = Function(W)




# Grab Currrent Time Before Running the Code
start = time.time()
solve(a == L, w, bcs)
# Grab Currrent Time After Running the Code
end = time.time()

#Subtract Start Time from The End Time
sim_time =[]
total_time = end - start
sim_time.append(total_time)

(u, P) = w.split()

#*****************************************
#  Pressure cross-section plot
#*****************************************

pressure= []
for i in np.linspace(0,1,11):
    point=(i,0.5)
    pressure.append(P(point))
print (pressure)

df = pd.DataFrame({'Cubic':pressure})



wb = load_workbook("/home/emmanuel/Documents/Manuscript_1/Codes/Homo/P_0.3   b_x/Horizontal/DFM.xlsx")

ws = wb['Sheet1']

for index, row in df.iterrows():
    cell = 'D%d'  % (index + 2)
    ws[cell] = row[0]
    


wb.save("/home/emmanuel/Documents/Manuscript_1/Codes/Homo/P_0.3   b_x/Horizontal/DFM.xlsx")


#defining output data files
vtkfileP = File('./output/Pressure.pvd')
vtkfileV = File('./output/Velocity.pvd')
vtkfilePhi = File('./output/Porosity.pvd')
vtkfileK = File('./output/Permeability.pvd')
    
plot(P)
vtkfileP << (P)
vtkfileV << (u)
vtkfilePhi << (poro)



#Subtract Start Time from The End Time
total_time = end - start
print("Total code execution time = "+ str(total_time) +"secs")


