import os
from fenics import *
from dolfin import *
from mshr import *
import numpy as np
import sys
import random
import math
import matplotlib.pyplot as plt

from openpyxl import load_workbook
import pandas as pd

from statistics import mean
# import the builtin time module
import time

#*****************************************#
               #CUBIC LAW
#*****************************************#         

def CL():
    #Initial and Boundary Condition
    u_inlet = 0.000167*100          #m/sec (1 m/sec = 6000 cm/min)
    
    #Rock and Fluid Properties
    mu = Constant(0.001)            #pa-sec (1 pa-sec = 1000 cp)
    poro_ref = 0.2                  #fraction
    perm_ref = 150E-15               #m2 (1 mD = 1E-15 m2)
    
    # Create a computational domain
    DimX = 1
    DimY = 1
    domain = Rectangle(Point(0,0), Point(DimX,DimY))
    
    # Create a subdomain for fractures (each frac is a rectangle)
    def createFracSubDomain (fracdata):
      fracdomain = []
    
      for i in range (0, len(fracdata)):
        L = fracdata[i][0]      # the length of your rectangle
        b = fracdata[i][1]      # aperture
        a = fracdata[i][2]      # orientation angle
        Cx = fracdata[i][3]     # center point x
        Cy = fracdata[i][4]     # center point y
    
        # coordinate of each corner point for rectangle with the center at (0,0)
        LeftTop_x = -L/2; LeftTop_y = b/2; LeftBot_x = -L/2; LeftBot_y = -b/2
        RightBot_x = L/2; RightBot_y = -b/2; RightTop_x = L/2; RightTop_y = b/2
    
        #The rotated position of each corner    
        Rx1 = Cx + (LeftTop_x  * math.cos(a)) - (LeftTop_y * math.sin(a))
        Ry1 = Cy + (LeftTop_x  * math.sin(a)) + (LeftTop_y * math.cos(a))
        Rx2 = Cx + (LeftBot_x  * math.cos(a)) - (LeftBot_y * math.sin(a))
        Ry2 = Cy + (LeftBot_x  * math.sin(a)) + (LeftBot_y * math.cos(a))
        Rx3 = Cx + (RightBot_x  * math.cos(a)) - (RightBot_y * math.sin(a))
        Ry3 = Cy + (RightBot_x  * math.sin(a)) + (RightBot_y * math.cos(a))
        Rx4 = Cx + (RightTop_x  * math.cos(a)) - (RightTop_y * math.sin(a))
        Ry4 = Cy + (RightTop_x  * math.sin(a)) + (RightTop_y * math.cos(a))
    
        domain_vertices = [Point(Rx1,Ry1),Point(Rx2,Ry2),Point(Rx3,Ry3),Point(Rx4,Ry4),Point(Rx1,Ry1)]
    
        if i == 0:
          fracdomain = Polygon(domain_vertices)
        else:
          fracdomain = fracdomain + Polygon(domain_vertices)
    
      return fracdomain
    
    FracWidth = 0.001
    # generate fracture subdmain based on fracture data 
    # fracdata = [length, aperture, angle, center_x, center_y]
    
    fracdata = [
            [0.3,FracWidth, -0.4,        0.2, 0.6],
            [0.4,FracWidth, 0.75,  0.3, 0.2],
            [0.3,FracWidth, 1,        0.7, 0.7],
            [0.5,FracWidth, 0.7,        0.4, 0.5],
            [0.2,FracWidth, 0.7,        0.4, 0.8],
            [0.5,FracWidth, -0.4,        0.4, 0.3],
            [0.3,FracWidth, -0.4,        0.6, 0.6],
            [0.2,FracWidth, 0.4,        0.8, 0.2],
            [0.3,FracWidth, -0.6,        0.3, 0.8],
            [0.3,FracWidth, -0.9,        0.8, 0.2],
            [0.1,FracWidth, -1,        0.9, 0.9]
            ]
    
    fracdomain = createFracSubDomain (fracdata)
    
    # Add fractures into the domain and mark them as subdomain '1'
    domain.set_subdomain(1, fracdomain)
    domain.set_subdomain(2, domain - fracdomain)
    
    # Generature mesh for the combined system
    mesh = generate_mesh(domain, 5)
    print("number of element", mesh.num_cells())
    
    
    # Create a cell list for fractures -> frac_cells
    domains = mesh.domains()
    subdomains = MeshFunction('size_t', mesh, 2, domains)
    frac_cells = SubsetIterator(subdomains, 1)
    dx = Measure('dx', domain=mesh, subdomain_data=subdomains)
    #plot(mesh)
    
    
    #Initialize Properties and Generate Cell Functions for Properties
    poro = MeshFunction("double", mesh, 2)
    perm = MeshFunction("double", mesh, 2)
    for cell in cells(mesh):
        poro[cell] = 0.175#poro_ref#random.gauss(poro_ref, 0.01)
        perm[cell] = perm_ref*pow(pow(poro[cell]/poro_ref,3)*pow((1-poro_ref)/(1-poro[cell]),2),6)
    
    #update properties in fracture cells
    for cell in frac_cells:
        poro[cell] = 0.9999999
        perm[cell] = FracWidth*FracWidth/12
    
    # Define Properties Functions and Initialize###############################
    print(max(poro.array()[:]),min(poro.array()[:]))
    print(max(perm.array()[:]),min(perm.array()[:]))
    
    property_code = """

    #include <pybind11/pybind11.h>
    #include <pybind11/eigen.h>
    namespace py = pybind11;
    
    #include <dolfin/function/Expression.h>
    #include <dolfin/mesh/MeshFunction.h>
    
    class Property : public dolfin::Expression
    {
    public:
    
      // Create expression with 2 components
      Property() : dolfin::Expression(2) {}
    
      // Function for evaluating expression on each cell
      void eval(Eigen::Ref<Eigen::VectorXd> values, Eigen::Ref<const Eigen::VectorXd> x, const ufc::cell& cell) const override
      {
        const uint cell_index = cell.index;
        values[0] = (*c0)[cell_index];
        values[1] = (*c1)[cell_index];
      }
    
      // The data stored in mesh functions
      std::shared_ptr<dolfin::MeshFunction<double>> c0;
      std::shared_ptr<dolfin::MeshFunction<double>> c1;
    
    };
    
    PYBIND11_MODULE(SIGNATURE, m)
    {
      py::class_<Property, std::shared_ptr<Property>, dolfin::Expression>
        (m, "Property")
        .def(py::init<>())
        .def_readwrite("c0", &Property::c0)
        .def_readwrite("c1", &Property::c1);
    }
    
    """
    
    c = CompiledExpression(compile_cpp_code(property_code).Property(),
                           c0=poro, c1=perm, degree=0)
    
    Phi = c[0]
    K = c[1]
       
    # Define function spaces
    #scalar = FiniteElement('P', triangle, 1)
    #vector = VectorElement('P', triangle, 2)
    
    scalar = FiniteElement('CG', triangle, 1)
    vector = VectorElement('CG', triangle, 2)
    
    system = FunctionSpace (mesh, MixedElement([vector, scalar]))
    
    # Create functions for boundary conditions
    noslip = Constant((0, 0))
    inflow = Constant((u_inlet, 0))
    zero   = Constant(0)
    
    def bd_noslip(x, on_boundary):
        return on_boundary and (x[1]>1-DOLFIN_EPS or x[1]<DOLFIN_EPS)                           
    def bd_inflow(x):
        return x[0]<DOLFIN_EPS #and x[1]<1-DOLFIN_EPS and x[1]>DOLFIN_EPS                             
    def bd_zero(x, on_boudary):
        return (x[0]>1-DOLFIN_EPS and x[1]<1-DOLFIN_EPS and x[1]>DOLFIN_EPS)
    
    # No-slip boundary condition for velocity
    bc0 = DirichletBC(system.sub(0), noslip, bd_noslip)
    # Inflow boundary condition for velocity
    bc1 = DirichletBC(system.sub(0), inflow, bd_inflow)
    # Boundary condition for pressure at outflow
    bc2 = DirichletBC(system.sub(1), zero, bd_zero)
    # Collect boundary conditions
    bcs = [bc0, bc1, bc2]#, bc2].
    
    # Define variational problem
    (v, q) = TestFunctions(system)
    (u, p) = TrialFunctions(system)
    
    f = Constant(0.0)
    
    a = (inner(u, v) + inner(grad(K/mu*p),v) + div(u)*q)*dx
    L = inner(q,f)*dx
    
    
    '''
    h = 2*Circumradius(mesh)
    beta  = 0.2
    delta = beta*h*h
    
    a1 = a1 + delta*inner(grad(q), grad(p))*dx(1)
    L1 = L1 + inner(delta*grad(q), f1)*dx(1)
    '''
    
    
    # Compute solution
    w = Function(system)
    
    # Grab Currrent Time Before Running the Code
    start = time.time()
    
    solve(a == L, w, bcs)
    
    # Grab Currrent Time After Running the Code
    end = time.time()
    
    #Subtract Start Time from The End Time
    sim_time =[]
    total_time = end - start
    sim_time.append(total_time)
    
    u, p = w.split()
    
    vol = assemble(Constant(1)*dx)
    avg_p= assemble(p*dx)/vol
    
    print(vol)
    print(avg_p)
    #*****************************************
    #  Pressure cross-section plot
    #*****************************************
    
    pressure= []
    for i in np.linspace(0,1,11):
        point=(i,0.5)
        pressure.append(p(point))
    print (pressure)
    
    df = pd.DataFrame({'Cubic':pressure})
    
    df_time = pd.DataFrame({'time':sim_time})
    
    wb = load_workbook("/home/emmanuel/Documents/Manuscript_1/Codes/multiple fracture/Errornorm_P_x b_x/Horizontal/DFM.xlsx")
    
    ws = wb['Sheet1']
    
    for index, row in df.iterrows():
        cell = 'B%d'  % (index + 2)
        ws[cell] = row[0]
        
    for index, row in df_time.iterrows():
        cell = 'B%d'  % (index + 13)
        ws[cell] = row[0]
    
    wb.save("/home/emmanuel/Documents/Manuscript_1/Codes/multiple fracture/Errornorm_P_x b_x/Horizontal/DFM.xlsx")
    
    
    #defining output data files
    vtkfileP = File('./outputCL/Pressure.pvd')
    vtkfileV = File('./outputCL/Velocity.pvd')
    vtkfilePhi = File('./outputCL/Porosity.pvd')
    vtkfileK = File('./outputCL/Permeability.pvd')
    
    plot(p)
    vtkfileP << (p)
    vtkfileV << (u)
    vtkfilePhi << (poro)  
    
    print("Total code execution time = "+ str(total_time) +"secs")
    
    return p, avg_p



#
##*****************************************************#
#          # Stokes Brinkman
##*****************************************************# 
#
def SB():

    #Initial and Boundary Condition
    u_inlet = 0.000167*100          #m/sec (1 m/sec = 6000 cm/min)
    
    #Rock and Fluid Properties
    mu = Constant(0.001)            #pa-sec (1 pa-sec = 1000 cp)
    poro_ref = 0.2                  #fraction
    perm_ref = 150E-15               #m2 (1 mD = 1E-15 m2)
    
    # Create a computational domain
    DimX = 1
    DimY = 1
    domain = Rectangle(Point(0,0), Point(DimX,DimY))
    
    # Create a subdomain for fractures (each frac is a rectangle)
    def createFracSubDomain (fracdata):
      fracdomain = []
    
      for i in range (0, len(fracdata)):
        L = fracdata[i][0]      # the length of your rectangle
        b = fracdata[i][1]      # aperture
        a = fracdata[i][2]      # orientation angle
        Cx = fracdata[i][3]     # center point x
        Cy = fracdata[i][4]     # center point y
    
        # coordinate of each corner point for rectangle with the center at (0,0)
        LeftTop_x = -L/2; LeftTop_y = b/2; LeftBot_x = -L/2; LeftBot_y = -b/2
        RightBot_x = L/2; RightBot_y = -b/2; RightTop_x = L/2; RightTop_y = b/2
    
        #The rotated position of each corner    
        Rx1 = Cx + (LeftTop_x  * math.cos(a)) - (LeftTop_y * math.sin(a))
        Ry1 = Cy + (LeftTop_x  * math.sin(a)) + (LeftTop_y * math.cos(a))
        Rx2 = Cx + (LeftBot_x  * math.cos(a)) - (LeftBot_y * math.sin(a))
        Ry2 = Cy + (LeftBot_x  * math.sin(a)) + (LeftBot_y * math.cos(a))
        Rx3 = Cx + (RightBot_x  * math.cos(a)) - (RightBot_y * math.sin(a))
        Ry3 = Cy + (RightBot_x  * math.sin(a)) + (RightBot_y * math.cos(a))
        Rx4 = Cx + (RightTop_x  * math.cos(a)) - (RightTop_y * math.sin(a))
        Ry4 = Cy + (RightTop_x  * math.sin(a)) + (RightTop_y * math.cos(a))
    
        domain_vertices = [Point(Rx1,Ry1),Point(Rx2,Ry2),Point(Rx3,Ry3),Point(Rx4,Ry4),Point(Rx1,Ry1)]
    
        if i == 0:
          fracdomain = Polygon(domain_vertices)
        else:
          fracdomain = fracdomain + Polygon(domain_vertices)
    
      return fracdomain
    
    FracWidth = 0.001
    # generate fracture subdmain based on fracture data 
    # fracdata = [length, aperture, angle, center_x, center_y]
    

      
    fracdata = [
            [0.3,FracWidth, -0.4,        0.2, 0.6],
            [0.4,FracWidth, 0.75,  0.3, 0.2],
            [0.3,FracWidth, 1,        0.7, 0.7],
            [0.5,FracWidth, 0.7,        0.4, 0.5],
            [0.2,FracWidth, 0.7,        0.4, 0.8],
            [0.5,FracWidth, -0.4,        0.4, 0.3],
            [0.3,FracWidth, -0.4,        0.6, 0.6],
            [0.2,FracWidth, 0.4,        0.8, 0.2],
            [0.3,FracWidth, -0.6,        0.3, 0.8],
            [0.3,FracWidth, -0.9,        0.8, 0.2],
            [0.1,FracWidth, -1,        0.9, 0.9]
            ]
    
    fracdomain = createFracSubDomain (fracdata)
    
    # Add fractures into the domain and mark them as subdomain '1'
    domain.set_subdomain(1, fracdomain)
    domain.set_subdomain(2, domain - fracdomain)
    
    # Generature mesh for the combined system
    mesh = generate_mesh(domain, 5)
    print("number of element", mesh.num_cells())
    
    # Create a cell list for fractures -> frac_cells
    domains = mesh.domains()
    subdomains = MeshFunction('size_t',mesh,2,domains)
    frac_cells = SubsetIterator(subdomains, 1)
    
    #plot(mesh)
    
    # Define function spaces and mixed (product) space
    scalar = FiniteElement("CG", mesh.ufl_cell(), 1)
    vector = VectorElement("CG", mesh.ufl_cell(), 2)
    W = FunctionSpace(mesh, MixedElement(vector, scalar))
    
    # Define trial and test functions
    (u, P) = TrialFunctions(W)
    (tf1, tf2) = TestFunctions(W)
    
    
    #Initialize Properties and Generate Cell Functions for Properties
    poro = MeshFunction("double", mesh, 2)
    perm = MeshFunction("double", mesh, 2)
    for cell in cells(mesh):
        poro[cell] = 0.175 #poro_ref#random.gauss(poro_ref, 0.01)
        perm[cell] = perm_ref*pow(pow(poro[cell]/poro_ref,3)*pow((1-poro_ref)/(1-poro[cell]),2),6)
    
    #update properties in fracture cells
    for cell in frac_cells:
        poro[cell] = 0.9999999
        perm[cell] = perm_ref*pow(pow(poro[cell]/poro_ref,3)*pow((1-poro_ref)/(1-poro[cell]),2),6)
    
    # Define Properties Functions and Initialize###############################
    #plot(poro)
    print(max(poro.array()[:]),min(poro.array()[:]))
    print(max(perm.array()[:]),min(perm.array()[:]))
    
    #******************************************************************
    property_code = """
    
    #include <pybind11/pybind11.h>
    #include <pybind11/eigen.h>
    namespace py = pybind11;
    
    #include <dolfin/function/Expression.h>
    #include <dolfin/mesh/MeshFunction.h>
    
    class Property2 : public dolfin::Expression
    {
    public:
    
      // Create expression with 2 components
      Property2() : dolfin::Expression(2) {}
    
      // Function for evaluating expression on each cell
      void eval(Eigen::Ref<Eigen::VectorXd> values, Eigen::Ref<const Eigen::VectorXd> x, const ufc::cell& cell) const override
      {
        const uint cell_index = cell.index;
        values[0] = (*c0)[cell_index];
        values[1] = (*c1)[cell_index];
      }
    
      // The data stored in mesh functions
      std::shared_ptr<dolfin::MeshFunction<double>> c0;
      std::shared_ptr<dolfin::MeshFunction<double>> c1;
    
    };
    
    PYBIND11_MODULE(SIGNATURE, m)
    {
      py::class_<Property2, std::shared_ptr<Property2>, dolfin::Expression>
        (m, "Property2")
        .def(py::init<>())
        .def_readwrite("c0", &Property2::c0)
        .def_readwrite("c1", &Property2::c1);
    }
    
    """
    
    c = CompiledExpression(compile_cpp_code(property_code).Property2(),
                           c0=poro, c1=perm, degree=0)
#    
    Phi = c[0]
    K = c[1]
    # Create functions for boundary conditions
    noslip = Constant((0, 0))
    inflow = Constant((u_inlet, 0))
    zero   = Constant(0)
    
    #Define boundary
    def boundary1(x):
        return x[0] < DOLFIN_EPS and x[1] > DOLFIN_EPS and x[1] < DimY - DOLFIN_EPS
    
    def boundary2(x):
        return (x[1] < DOLFIN_EPS or x[1] > DimY - DOLFIN_EPS) #or (x[0] < DOLFIN_EPS and x[1] <= 0.48) or (x[0] < DOLFIN_EPS and x[1] >= 1-0.48)
    
    def boundary3(x):
        return x[0] > DimX - DOLFIN_EPS
    
    # Define essential boundary
    # No-slip boundary condition for velocity
    bc0 = DirichletBC(W.sub(0), noslip, boundary2)
    # Inflow boundary condition for velocity
    bc1 = DirichletBC(W.sub(0), inflow, boundary1)
    # Boundary condition for pressure at outflow
    bc2 = DirichletBC(W.sub(1), zero, boundary3)
    # Collect boundary conditions
    bcs = [bc0, bc1, bc2]
    
    
    f = Constant(0.0)
    
    a = mu/Phi*inner(grad(u),grad(tf1))*dx + inner(grad(P),tf1)*dx + mu/K*inner(u,tf1)*dx + div(u)*tf2*dx
    L = dot(f,tf2)*dx
    
    w = Function(W)
    
    
    
    
    # Grab Currrent Time Before Running the Code
    start = time.time()
    solve(a == L, w, bcs)
    # Grab Currrent Time After Running the Code
    end = time.time()
    
    #Subtract Start Time from The End Time
    sim_time =[]
    total_time = end - start
    sim_time.append(total_time)
    
    (u, P) = w.split()
    
#    vol = assemble(Constant(1)*dx)
#    avg_P= assemble(P*dx)/vol
    
#    print(vol)
#    print(avg_P)    
    #*****************************************
    #  Pressure cross-section plot
    #*****************************************
    
    pressure = []
    for i in np.linspace(0,1,11):
        point=(i,0.5)
        pressure.append(P(point))
    print (pressure)
    
    df = pd.DataFrame({'StokesBrinkman':pressure})
    
    df_time = pd.DataFrame({'time':sim_time})
    
    wb = load_workbook("/home/emmanuel/Documents/Manuscript_1/Codes/multiple fracture/Errornorm_P_x b_x/Horizontal/DFM.xlsx")
    
    ws = wb['Sheet1']
    
    for index, row in df.iterrows():
        cell = 'D%d'  % (index + 2)
        ws[cell] = row[0]
        
    for index, row in df_time.iterrows():
        cell = 'D%d'  % (index + 13)
        ws[cell] = row[0]    
    
    wb.save("/home/emmanuel/Documents/Manuscript_1/Codes/multiple fracture/Errornorm_P_x b_x/Horizontal/DFM.xlsx")
    
    
    #defining output data files
    vtkfileP = File('./outputSB/Pressure.pvd')
    vtkfileV = File('./outputSB/Velocity.pvd')
    vtkfilePhi = File('./outputSB/Porosity.pvd')
    vtkfileK = File('./outputSB/Permeability.pvd')
        
    plot(P)
    vtkfileP << (P)
    vtkfileV << (u)
    vtkfilePhi << (poro)
    
    
    
    #Subtract Start Time from The End Time
    total_time = end - start
    print("Total code execution time = "+ str(total_time) +"secs")
    
    return P#, avg_P


# Error function
Cl,Cl_avg= CL()
Sb = SB() #, Sb_avg 
 
E = errornorm(Cl, Sb, 'L2')

print(E)


#CL()
#SB()

## compute maximum error at vertices
#vertex_CL = CL().compute_vertex_values(mesh)
#vertex_SB = SB().compute_vertex_values(mesh)
#error_max = np.max(np.abs(vertex_CL - vertex_SB))

# Create a computational domain
DimX = 1
DimY = 1
domain = Rectangle(Point(0,0), Point(DimX,DimY))

# Create a subdomain for fractures (each frac is a rectangle)
def createFracSubDomain (fracdata):
  fracdomain = []

  for i in range (0, len(fracdata)):
    L = fracdata[i][0]      # the length of your rectangle
    b = fracdata[i][1]      # aperture
    a = fracdata[i][2]      # orientation angle
    Cx = fracdata[i][3]     # center point x
    Cy = fracdata[i][4]     # center point y

    # coordinate of each corner point for rectangle with the center at (0,0)
    LeftTop_x = -L/2; LeftTop_y = b/2; LeftBot_x = -L/2; LeftBot_y = -b/2
    RightBot_x = L/2; RightBot_y = -b/2; RightTop_x = L/2; RightTop_y = b/2

    #The rotated position of each corner    
    Rx1 = Cx + (LeftTop_x  * math.cos(a)) - (LeftTop_y * math.sin(a))
    Ry1 = Cy + (LeftTop_x  * math.sin(a)) + (LeftTop_y * math.cos(a))
    Rx2 = Cx + (LeftBot_x  * math.cos(a)) - (LeftBot_y * math.sin(a))
    Ry2 = Cy + (LeftBot_x  * math.sin(a)) + (LeftBot_y * math.cos(a))
    Rx3 = Cx + (RightBot_x  * math.cos(a)) - (RightBot_y * math.sin(a))
    Ry3 = Cy + (RightBot_x  * math.sin(a)) + (RightBot_y * math.cos(a))
    Rx4 = Cx + (RightTop_x  * math.cos(a)) - (RightTop_y * math.sin(a))
    Ry4 = Cy + (RightTop_x  * math.sin(a)) + (RightTop_y * math.cos(a))

    domain_vertices = [Point(Rx1,Ry1),Point(Rx2,Ry2),Point(Rx3,Ry3),Point(Rx4,Ry4),Point(Rx1,Ry1)]

    if i == 0:
      fracdomain = Polygon(domain_vertices)
    else:
      fracdomain = fracdomain + Polygon(domain_vertices)

  return fracdomain

FracWidth = 0.001
# generate fracture subdmain based on fracture data 
# fracdata = [length, aperture, angle, center_x, center_y]


  
fracdata = [
            [0.3,FracWidth, -0.4,        0.2, 0.6],
            [0.4,FracWidth, 0.75,  0.3, 0.2],
            [0.3,FracWidth, 1,        0.7, 0.7],
            [0.5,FracWidth, 0.7,        0.4, 0.5],
            [0.2,FracWidth, 0.7,        0.4, 0.8],
            [0.5,FracWidth, -0.4,        0.4, 0.3],
            [0.3,FracWidth, -0.4,        0.6, 0.6],
            [0.2,FracWidth, 0.4,        0.8, 0.2],
            [0.3,FracWidth, -0.6,        0.3, 0.8],
            [0.3,FracWidth, -0.9,        0.8, 0.2],
            [0.1,FracWidth, -1,        0.9, 0.9]
            ]

fracdomain = createFracSubDomain (fracdata)

# Add fractures into the domain and mark them as subdomain '1'
domain.set_subdomain(1, fracdomain)
domain.set_subdomain(2, domain - fracdomain)

# Generature mesh for the combined system
mesh = generate_mesh(domain, 5)

W = FunctionSpace (mesh, "CG", 1)

#print(Cl.vector()[:]-Sb.vector()[:])

error = project(abs(Cl-Sb), W)
#defining output data files
vtkfileerror = File('./outputerror/Error.pvd')
# 
plot(error)
vtkfileerror  << (error)  






          
